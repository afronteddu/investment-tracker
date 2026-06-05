"""
Parse DeGiro transaction exports → compute current positions with avg cost basis.
Drop new exports into data/transactions/ and this recomputes automatically.
"""
from __future__ import annotations

import glob
import os
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl

# ISIN → Yahoo Finance ticker mapping
ISIN_TO_TICKER = {
    "US0381692070": "APLD",       # Applied Digital Corp
    "US92537N1081": "VRT",        # Vertiv Holdings
    "NL0009805522": "NBIS",       # Nebius Group (ex-Yandex)
    "US80004C2008": "SNDK",       # SanDisk Corp
    "US67066G1040": "NVDA",       # Nvidia
    "NL0010273215": "ASML.AS",    # ASML (Euronext Amsterdam)
    "US0846707026": "BRK-B",      # Berkshire Hathaway B
    "IE00B4ND3602": "IGLN.L",     # iShares Physical Gold ETC (LSE, USD)
    "IE00B4NCWG09": "ISLN.L",     # iShares Physical Silver ETC (LSE, USD)
    "IE00B4L5Y983": "IWDA.AS",    # iShares Core MSCI World
    "IE00B3XXRP09": "VUSA.AS",    # Vanguard S&P 500
    "IE00B4K48X80": "IMAE.AS",    # iShares Core MSCI Europe
    "LU2009202107": "AEME.PA",    # Amundi MSCI Emerging ex-China
    "IE00B4WXJJ64": "IEAG.AS",    # iShares Core EUR Govt Bond
}

TICKER_NAMES = {
    "APLD": "Applied Digital Corp",
    "VRT": "Vertiv Holdings",
    "NBIS": "Nebius Group",
    "SNDK": "SanDisk Corp",
    "NVDA": "Nvidia",
    "ASML.AS": "ASML Holding NV",
    "BRK-B": "Berkshire Hathaway B",
    "IGLN.L": "iShares Physical Gold ETC",
    "ISLN.L": "iShares Physical Silver ETC",
    "IWDA.AS": "iShares MSCI World",
    "VUSA.AS": "Vanguard S&P 500",
    "IMAE.AS": "iShares MSCI Europe",
    "AEME.PA": "Amundi MSCI EM ex-China",
    "IEAG.AS": "iShares EUR Govt Bond",
    # Watchlist & hot picks universe
    "SMCI": "Super Micro Computer",
    "AMD": "Advanced Micro Devices",
    "PLTR": "Palantir Technologies",
    "IONQ": "IonQ",
    "RGTI": "Rigetti Computing",
    "QUBT": "Quantum Computing Inc",
    "QBTS": "D-Wave Quantum",
    "TSM": "Taiwan Semiconductor",
    "ARM": "Arm Holdings",
    "SPY": "S&P 500 ETF",
    "QQQ": "Nasdaq 100 ETF",
    "VWRL.AS": "Vanguard FTSE All-World",
    "MSFT": "Microsoft",
    "GOOGL": "Alphabet (Google)",
    "META": "Meta Platforms",
    "AMZN": "Amazon",
    "TSLA": "Tesla",
    "AAPL": "Apple",
    "AVGO": "Broadcom",
    "MRVL": "Marvell Technology",
    "INTC": "Intel",
    "QCOM": "Qualcomm",
    "AI": "C3.ai",
    "SOUN": "SoundHound AI",
    "BBAI": "BigBear.ai",
    "UPST": "Upstart Holdings",
    "PATH": "UiPath",
    "SNOW": "Snowflake",
    "VST": "Vistra Corp",
    "CEG": "Constellation Energy",
    "NRG": "NRG Energy",
    "RXRX": "Recursion Pharmaceuticals",
    "TMDX": "TransMedics Group",
    "ARKK": "ARK Innovation ETF",
    "BOTZ": "Global X Robotics & AI ETF",
    "AIQ": "Global X AI & Technology ETF",
    # House savings CORE
    "UCG.MI": "UniCredit SpA",
    "NOVN.SW": "Novartis AG",
    "ENEL.MI": "Enel SpA",
    "AXA.PA": "AXA SA",
    "IBE.MC": "Iberdrola SA",
    # House savings SATELLITE
    "TTE.PA": "TotalEnergies SE",
    "GSK.L": "GSK plc",
    # House savings GROWTH
    "ROG.SW": "Roche Holding AG",
    "PYPL": "PayPal Holdings",
    "ABNB": "Airbnb Inc",
    # Defence / 2027 HC watch
    "LDO.MI": "Leonardo SpA",
    "RHM.DE": "Rheinmetall AG",
}

BUCKET_MAP = {
    "APLD": "high_conviction",
    "VRT": "high_conviction",
    "NBIS": "high_conviction",
    "SNDK": "growth",
    "NVDA": "growth",
    "ASML.AS": "growth",
    "BRK-B": "retirement",
    "IGLN.L": "retirement",
    "ISLN.L": "retirement",
    "IWDA.AS": "retirement",
    "VUSA.AS": "retirement",
    "IMAE.AS": "retirement",
    "AEME.PA": "retirement",
    "IEAG.AS": "retirement",
}


@dataclass
class Position:
    ticker: str
    name: str
    isin: str
    shares: float = 0.0
    total_cost_eur: float = 0.0   # total EUR spent (including fees)
    buy_count: int = 0
    first_buy_date: str = ""      # YYYY-MM-DD of earliest buy

    @property
    def avg_cost_eur(self) -> float:
        return self.total_cost_eur / self.shares if self.shares > 0 else 0.0

    @property
    def bucket(self) -> str:
        return BUCKET_MAP.get(self.ticker, "growth")

    def to_dict(self) -> dict:
        return {
            "ticker": self.ticker,
            "name": self.name,
            "isin": self.isin,
            "shares": round(self.shares, 4),
            "avg_cost_eur": round(self.avg_cost_eur, 4),
            "total_cost_eur": round(self.total_cost_eur, 2),
            "buy_count": self.buy_count,
            "bucket": self.bucket,
            "first_buy_date": self.first_buy_date,
        }


def load_transactions(data_dir: str = "data/transactions") -> list[dict]:
    """Load and merge all transaction xlsx files, deduplicate by order ID."""
    rows = []
    seen_orders = set()

    pattern = os.path.join(data_dir, "*.xlsx")
    files = sorted(glob.glob(pattern))

    for filepath in files:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = row
                continue
            if not row[0]:
                continue
            # Columns: Data, Ora, Prodotto, ISIN, Borsa rif, Borsa, Quantità,
            #          Quotazione, currency, Valore locale, currency, Valore EUR,
            #          Tasso cambio, AutoFX, Costi EUR, Totale EUR, ID Ordine, ...
            order_id = row[16] if len(row) > 16 else None

            # Deduplicate: prefer order_id when present, otherwise use content key
            dedup_key = order_id if order_id else (row[0], row[3], str(row[6]), str(row[15]))
            if dedup_key in seen_orders:
                continue
            seen_orders.add(dedup_key)

            rows.append({
                "date": row[0],
                "time": row[1],
                "product": row[2],
                "isin": row[3],
                "quantity": row[6],
                "price": row[7],
                "price_currency": row[8],
                "value_local": row[9],
                "value_eur": row[11],
                "fx_rate": row[12],
                "autofx_cost": row[13],
                "transaction_cost_eur": row[14],
                "total_eur": row[15],
                "order_id": order_id,
            })

    return rows


def compute_positions(data_dir: str = "data/transactions") -> dict[str, Position]:
    transactions = load_transactions(data_dir)

    # Must process chronologically so sells never arrive before their buys
    def parse_date(tx):
        try:
            return datetime.strptime(tx["date"], "%d-%m-%Y")
        except Exception:
            return datetime.min

    transactions.sort(key=parse_date)

    positions: dict[str, Position] = {}

    for tx in transactions:
        isin = tx.get("isin")
        qty = tx.get("quantity")
        total_eur = tx.get("total_eur")

        if not isin or not qty or not total_eur:
            continue

        # Skip sells (positive total = money in = sell)
        # Buys have negative total_eur (money out)
        if total_eur > 0:
            # This is a sell — reduce position
            ticker = ISIN_TO_TICKER.get(isin, isin)
            if ticker in positions and positions[ticker].shares > 0:
                sell_frac = abs(qty) / positions[ticker].shares
                positions[ticker].total_cost_eur *= (1 - sell_frac)
                positions[ticker].shares -= abs(qty)
            continue

        ticker = ISIN_TO_TICKER.get(isin, isin)
        product = tx.get("product", ticker)

        if ticker not in positions:
            positions[ticker] = Position(ticker=ticker, name=product, isin=isin)

        positions[ticker].shares += qty
        positions[ticker].total_cost_eur += abs(total_eur)
        positions[ticker].buy_count += 1
        if not positions[ticker].first_buy_date:
            try:
                positions[ticker].first_buy_date = datetime.strptime(tx["date"], "%d-%m-%Y").strftime("%Y-%m-%d")
            except Exception:
                pass

    # Drop positions with 0 shares (fully sold)
    return {k: v for k, v in positions.items() if v.shares > 0.001}


def compute_closed_positions(data_dir: str = "data/transactions") -> list[dict]:
    """Return metadata for fully-exited positions: avg cost, first buy, last sell date."""
    transactions = load_transactions(data_dir)

    def parse_date(tx):
        try:
            return datetime.strptime(tx["date"], "%d-%m-%Y")
        except Exception:
            return datetime.min

    transactions.sort(key=parse_date)

    # Track buys and sells separately per ticker
    ticker_info: dict[str, dict] = {}

    for tx in transactions:
        isin = tx.get("isin")
        qty = tx.get("quantity")
        total_eur = tx.get("total_eur")
        if not isin or not qty or not total_eur:
            continue

        ticker = ISIN_TO_TICKER.get(isin, isin)
        if ticker not in ticker_info:
            ticker_info[ticker] = {
                "ticker": ticker,
                "name": TICKER_NAMES.get(ticker, ticker),
                "isin": isin,
                "shares": 0.0,
                "total_cost_eur": 0.0,
                "first_buy_date": None,
                "last_sell_date": None,
                "bucket": BUCKET_MAP.get(ticker, "retirement"),
            }
        info = ticker_info[ticker]

        d = parse_date(tx).strftime("%Y-%m-%d")
        if total_eur < 0:
            # Buy
            info["shares"] += qty
            info["total_cost_eur"] += abs(total_eur)
            if info["first_buy_date"] is None:
                info["first_buy_date"] = d
            # Track peak shares (max ever held)
            if info["shares"] > info.get("peak_shares", 0):
                info["peak_shares"] = info["shares"]
        else:
            # Sell
            info["shares"] -= abs(qty)
            info["last_sell_date"] = d

    # Only return positions that are fully exited
    return [
        v for v in ticker_info.values()
        if abs(v["shares"]) < 0.01 and v["last_sell_date"] is not None
    ]


def compute_lifetime_stats(data_dir: str = "data/transactions") -> dict:
    """Compute all-time realized P&L, total deployed, and monthly flows from raw transactions."""
    import openpyxl
    from collections import defaultdict

    holdings: dict = defaultdict(lambda: {'shares': 0.0, 'cost': 0.0})
    realized_pnl = 0.0
    total_deployed = 0.0
    total_returned = 0.0
    monthly: dict = defaultdict(float)

    seen: set = set()
    files = sorted(glob.glob(os.path.join(data_dir, "*.xlsx")))
    rows = []
    for filepath in files:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row[0]:
                continue
            key = (row[0], row[3], str(row[6]), str(row[15]))
            if key in seen:
                continue
            seen.add(key)
            rows.append({'date': row[0], 'product': row[2], 'qty': float(row[6] or 0), 'total_eur': float(row[15] or 0)})

    rows.sort(key=lambda x: datetime.strptime(x['date'], '%d-%m-%Y'))

    for r in rows:
        prod = r['product']
        t = r['total_eur']
        qty = r['qty']
        month = datetime.strptime(r['date'], '%d-%m-%Y').strftime('%Y-%m')
        if t < 0:
            total_deployed += abs(t)
            monthly[month] += abs(t)
            holdings[prod]['shares'] += qty
            holdings[prod]['cost'] += abs(t)
        else:
            total_returned += t
            monthly[month] -= t
            if holdings[prod]['shares'] > 0:
                frac = abs(qty) / holdings[prod]['shares']
                realized_pnl += t - holdings[prod]['cost'] * frac
                holdings[prod]['cost'] *= (1 - frac)
                holdings[prod]['shares'] -= abs(qty)

    return {
        "total_deployed": round(total_deployed, 2),
        "total_returned": round(total_returned, 2),
        "realized_pnl": round(realized_pnl, 2),
        "monthly_flows": [{"month": m, "net": round(v, 2)} for m, v in sorted(monthly.items())],
    }


def add_new_export(src_path: str, data_dir: str = "data/transactions"):
    """Copy a new DeGiro export into the data directory."""
    import shutil
    filename = os.path.basename(src_path)
    dest = os.path.join(data_dir, filename)
    shutil.copy2(src_path, dest)
    return dest
